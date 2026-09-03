"""
Prices, quotes and the Xero hand-off.

A price is looked up in this order, and the first answer wins:

    1. The part number, in the imported price list. FPFG425-040 has a price;
       that is the price, whatever anything else says.
    2. A rate per square metre for that filter type and media, if one is set.
       0.20 m2 of G4 flat panel at $34/m2 is $6.80.
    3. Nothing. The line is quoted at zero and flagged, never guessed at —
       a made-up price on a quote goes out to a customer under TAF's name.

Reading price spreadsheets is deliberately forgiving about column names,
because the files come from several places and none of them agree on
"Part No" vs "Code" vs "SKU".
"""
from __future__ import annotations

import csv
import datetime as _dt
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from . import part_numbers as _pn

# ── Reading a price file ─────────────────────────────────────────────────────

CODE_HEADINGS = ("part number", "part no", "partno", "part", "code",
                 "item code", "itemcode", "sku", "product code", "item",
                 "product", "stock code", "catalogue number")
PRICE_HEADINGS = ("price", "unit price", "sell", "sell price", "rate",
                  "amount", "each", "unit amount", "list price", "ex gst",
                  "unit cost", "cost")
DESC_HEADINGS = ("description", "desc", "details", "name", "product name")

# The description that goes on a quote and an invoice is the SALES one, for
# the same reason as the price: a Xero item export carries a purchases
# description too, and it is not what a customer should read.
DESC_PREFERRED = ("sales", "sell")
DESC_AVOIDED = ("purchase", "purchases", "cost", "supplier")

# The item's own name, kept alongside the sales description. In a Xero export
# the sales description is a template — "Flat Panel Filter G4 Rating / Size:"
# on every row, with the size left to be filled in — so it is the wrong thing
# to show in a product list, while the item name says which product it is.
NAME_HEADINGS = ("item name", "itemname", "name", "product name", "title")

# What TAF charges, not what TAF pays. A Xero item export carries both a
# PurchasesUnitPrice and a SalesUnitPrice; picking the first column whose
# heading merely contains "price" lands on the purchase column, which in an
# export like that is empty — so every row is dropped and the import looks
# like it found nothing. Headings are scored instead of taken in order.
PRICE_PREFERRED = ("sales", "sell", "selling", "list", "retail", "charge")
PRICE_AVOIDED = ("purchase", "purchases", "cost", "buy", "supplier", "wholesale")


# A column heading is a label, not a sentence. Without this cap a Read Me
# paragraph that happens to contain the words "code", "price" and
# "description" is read as a header row.
MAX_HEADING_LENGTH = 40


def _norm_heading(text: Any) -> str:
    cleaned = re.sub(r"[^a-z0-9 ]+", " ", str(text or "").lower()).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned if len(cleaned) <= MAX_HEADING_LENGTH else ""


def _money(value: Any) -> Optional[float]:
    """A number from a cell that may be '$12.50', '12.50 ea' or 12.5."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if not text or text in ("-", "."):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _score(cell: str, exact: tuple, prefer: tuple = (), avoid: tuple = ()) -> int:
    """How well a heading matches, or 0 for not at all. Higher is better."""
    if not cell:
        return 0
    base = 0
    if cell in exact:
        base = 100
    elif any(h in cell for h in exact):
        base = 50
    if not base:
        return 0
    if any(w in cell for w in prefer):
        base += 30
    if any(w in cell for w in avoid):
        base -= 45
    return max(base, 1)


def _candidates(row: Iterable[Any]) -> Dict[str, List[int]]:
    """Every column a header row offers for each field, best-scoring first."""
    cells = [_norm_heading(c) for c in row]
    out: Dict[str, List[int]] = {}
    for field, exact, prefer, avoid in (
            ("code",  CODE_HEADINGS,  (), ()),
            ("price", PRICE_HEADINGS, PRICE_PREFERRED, PRICE_AVOIDED),
            ("desc",  DESC_HEADINGS,  DESC_PREFERRED, DESC_AVOIDED),
            ("name",  NAME_HEADINGS,  (), DESC_AVOIDED)):
        scored = [(_score(c, exact, prefer, avoid), i)
                  for i, c in enumerate(cells)]
        # Best score first, and among equal scores the leftmost column. The
        # score is what decides — leaving ties to fall out of the sort order
        # once meant the right column was picked by accident, which is not
        # something to build a price list on.
        ranked = sorted((si for si in scored if si[0] > 0),
                        key=lambda si: (-si[0], si[1]))
        if ranked:
            out[field] = [i for _s, i in ranked]
    return out


def _pick_columns(row: Iterable[Any]) -> Optional[Dict[str, int]]:
    """Which columns of a header row hold the code, price and description.

    The best-scoring candidate for each. Where several columns could be the
    price, ``_best_price_column`` re-checks the choice against the rows that
    follow — a heading can look right and still be an empty column.
    """
    cands = _candidates(row)
    if "code" not in cands or "price" not in cands:
        return None
    return {field: idxs[0] for field, idxs in cands.items()}


def _sheets_from_csv(path: Path) -> List[Tuple[str, List[List[Any]]]]:
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return [(path.name, [row for row in csv.reader(f)])]


def _sheets_from_excel(path: Path) -> List[Tuple[str, List[List[Any]]]]:
    """Each worksheet separately — a heading found on one sheet must not be
    used to read another. These workbooks carry a Read Me and a working
    pricing model alongside the catalogue, and both would otherwise be read
    as if they had the catalogue's columns."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: List[Tuple[str, List[List[Any]]]] = []
    try:
        for ws in wb.worksheets:
            sheets.append((ws.title,
                           [list(row) for row in ws.iter_rows(values_only=True)]))
    finally:
        wb.close()
    return sheets


def _best_price_column(rows: List[List[Any]], start: int,
                       code_col: int, price_cols: List[int]) -> int:
    """Of the columns that could be the price, the one that actually is.

    A heading can score well and hold nothing — a Xero item export has both a
    purchases and a sales unit price, and one of them is usually blank. The
    column carrying real numbers next to real codes wins; ties go to the
    better-scoring heading, which is the order they arrive in.
    """
    if len(price_cols) < 2:
        return price_cols[0]
    filled = {c: 0 for c in price_cols}
    for row in rows[start:start + 400]:
        if code_col >= len(row) or not str(row[code_col] or "").strip():
            continue
        for c in price_cols:
            if c < len(row) and _money(row[c]) is not None:
                filled[c] += 1
    best = max(price_cols, key=lambda c: filled[c])
    return best if filled[best] else price_cols[0]


def _heading_of(row: List[Any], idx: int) -> str:
    """The heading text at a column, for the report."""
    if idx is None or idx >= len(row):
        return "?"
    return str(row[idx] or "?").strip() or "?"


def read_price_file(path) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    """Read one price spreadsheet. Returns (rows, problems, report).

    Every sheet in a workbook is read, and the header row is found rather
    than assumed to be the first — these files usually open with a title and
    a blank line or two. `report` names the columns each sheet was read from.
    """
    path = Path(path)
    problems: List[str] = []
    report: List[str] = []
    try:
        sheets = (_sheets_from_csv(path) if path.suffix.lower() in (".csv", ".txt")
                  else _sheets_from_excel(path))
    except Exception as exc:
        return [], [f"{path.name}: could not be opened ({exc})"], []

    out: List[Dict[str, Any]] = []
    for sheet_name, raw in sheets:
        cols: Optional[Dict[str, int]] = None
        for n, row in enumerate(raw):
            if not row or all(c in (None, "") for c in row):
                continue
            if cols is None:
                cands = _candidates(row)
                if "code" in cands and "price" in cands:
                    cols = {f: idxs[0] for f, idxs in cands.items()}
                    cols["price"] = _best_price_column(
                        raw, n + 1, cols["code"], cands["price"])
                    # Say which columns were used. Reading the wrong one is
                    # the failure that looks like "it imported nothing", and
                    # it should never again be something you can't see.
                    report.append(f"{path.name} · {sheet_name}: " + ", ".join(
                        f"{f} = {_heading_of(row, cols[f])}"
                        for f in ("code", "price", "desc", "name") if f in cols))
                continue          # the heading row itself is never data

            def _cell(name, _row=row, _cols=cols):
                i = _cols.get(name)
                return _row[i] if i is not None and i < len(_row) else None

            code = str(_cell("code") or "").strip()
            price = _money(_cell("price"))
            if not code or price is None:
                continue
            if not re.search(r"[0-9]", code):
                continue          # a section heading, not a part number
            out.append({
                "part_number": code.upper(),
                "description": str(_cell("desc") or "").strip(),
                "name":        str(_cell("name") or "").strip(),
                "unit_price":  round(price, 4),
                "sheet":       sheet_name,
            })

    if not out and not problems:
        problems.append(
            f"{path.name}: no priced rows found — the sheet needs a heading "
            "row with a part-number column and a price column.")
    # Later rows win, so a corrected price further down the file is the one kept.
    deduped: Dict[str, Dict[str, Any]] = {}
    for row in out:
        deduped[row["part_number"]] = row
    return list(deduped.values()), problems, report


def read_price_files(paths) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    """Read several price files into one list, later files winning."""
    merged: Dict[str, Dict[str, Any]] = {}
    problems: List[str] = []
    report: List[str] = []
    for p in paths:
        rows, probs, rep = read_price_file(p)
        problems.extend(probs)
        report.extend(rep)
        for row in rows:
            merged[row["part_number"]] = row
    return list(merged.values()), problems, report


# ── Pricing a line ───────────────────────────────────────────────────────────

def _rate_key(filter_type: str, media: str) -> str:
    return f"{_pn.wording_key(filter_type)}|{_pn.wording_key(media)}"


# How far a thickness may be rounded to find a price. The catalogue is priced
# in 5mm steps, but purchase orders ask for 48mm and 47mm — sizes that are
# made out of the same materials as the 50mm and cost the same to make.
# Anything further away than this is a different filter, not a near miss.
NEAREST_THICKNESS_MM = 4


def nearest_priced_part(item: Dict[str, Any],
                        prices: Dict[str, float]) -> Tuple[str, float]:
    """A priced part number for a thickness the list doesn't carry.

    A 48mm V-form is not in a catalogue priced in 5mm steps, but it is the
    same filter as the 50mm to within a couple of millimetres. Returns the
    part number found and its price, or ("", 0.0).

    Deliberately narrow: same media, same area, same type — only the
    thickness moves, and only by a few millimetres. And the caller is told
    this is what happened, so it never reads as a listed price.
    """
    ftype = (item.get("Filter Type") or "").strip().lower()
    prefix = PART_PREFIX_FOR_NEAREST.get(ftype)
    if not prefix or not prices:
        return ("", 0.0)
    try:
        thickness = int(str(item.get("Channel") or "").strip())
    except (TypeError, ValueError):
        return ("", 0.0)
    if thickness <= 0:
        return ("", 0.0)

    code = _pn.media_code(item.get("Media Type") or "")
    nominal = _pn.nominal_square_metres(_pn.effective_area(item))
    if not code or nominal <= 0:
        return ("", 0.0)
    suffix = _pn.sqm_suffix(nominal)

    # Nearest first, so 48 tries 50 before 45 when both are priced.
    for step in sorted(range(-NEAREST_THICKNESS_MM, NEAREST_THICKNESS_MM + 1),
                       key=lambda d: (abs(d), -d)):
        if step == 0:
            continue
        candidate = f"{prefix}{code}{thickness + step}-{suffix}"
        listed = prices.get(candidate)
        if listed is not None:
            return (candidate, round(float(listed), 4))
    return ("", 0.0)


PART_PREFIX_FOR_NEAREST = dict(_pn.PART_NUMBER_PREFIXES)


def price_for_item(item: Dict[str, Any],
                   prices: Optional[Dict[str, float]] = None,
                   rates: Optional[Dict[str, float]] = None) -> Tuple[float, str]:
    """The unit price for one line, and where it came from.

    Returns ``(unit_price, source)`` where source is "list", "near", "rate"
    or "" — an empty source means nothing priced it and the line needs a
    human. "near" is a price taken from the same filter a few millimetres
    thicker or thinner, and is always shown as such.
    """
    # A line someone entered a price on themselves — picked out of the price
    # list, or a one-off nobody has a part number for. It is already priced,
    # and guessing at it from dimensions it doesn't have would be worse.
    if item.get("item_kind") == "catalogue":
        try:
            unit = float(item.get("Unit Price") or 0)
        except (TypeError, ValueError):
            unit = 0.0
        if unit > 0:
            return (round(unit, 4), item.get("_price_source") or "list")
        # Fall through: it may still carry a part number the list knows.

    part = (item.get("Part Number") or "").strip().upper()
    if part and prices:
        listed = prices.get(part)
        if listed is not None:
            return (round(float(listed), 4), "list")
        near_part, near_price = nearest_priced_part(item, prices)
        if near_part:
            item["_priced_as"] = near_part      # shown, never silently used
            return (near_price, "near")

    if rates:
        ftype = (item.get("Filter Type") or "").strip()
        media = (item.get("Media Type") or "").strip()
        rate = rates.get(_rate_key(ftype, media))
        if rate is None:
            rate = rates.get(_rate_key(ftype, ""))   # any media
        if rate is not None:
            area = _pn.nominal_square_metres(_pn.effective_area(item))
            if area > 0:
                return (round(float(rate) * area, 4), "rate")
    return (0.0, "")


def why_unpriced(item: Dict[str, Any],
                 prices: Optional[Dict[str, float]] = None) -> str:
    """Why a line found no price — a gap in the list, or a bad line.

    The difference matters: "this size isn't in the price list" is something
    to go and add, while "this line has no part number" is something wrong
    with the line itself. Guessing a nearby price instead would hide both.
    """
    # A line typed in by hand or picked from the list carries its own price.
    # If it has none, that is simply a price nobody has entered — nothing to
    # do with filter types or dimensions it was never going to have.
    if item.get("item_kind") == "catalogue":
        part = (item.get("Part Number") or "").strip().upper()
        if part and prices and part not in prices:
            return f"{part} is not in the price list"
        return "no price entered for this line"

    part = (item.get("Part Number") or "").strip().upper()
    if not part:
        ftype = (item.get("Filter Type") or "").strip()
        if not ftype:
            return "the line has no filter type, so it has no part number"
        return "the line has no part number — check its size and media"
    if not prices:
        return "no price list has been imported"

    # Other areas of the same product tell us whether the product is priced
    # at all, or just not in this size.
    stem, _, suffix = part.rpartition("-")
    same = sorted({_suffix_area(p.rpartition("-")[2])
                   for p in prices if p.rpartition("-")[0] == stem}
                  - {None})
    if same:
        return (f"priced from {same[0]:.1f} to {same[-1]:.1f} m2 — "
                f"not at {(_suffix_area(suffix) or 0):.1f} m2")

    ftype = (item.get("Filter Type") or "").strip()
    if ftype.lower() in _pn.FIXED_PART_NUMBER_STEMS:
        # A stepped filter or flyscreen: its media and thickness are fixed and
        # not what the code is built from, so naming them would only mislead.
        return f"{ftype.lower()}s are not in the price list"
    media = (item.get("Media Type") or "").strip() or "that media"
    channel = str(item.get("Channel") or "").strip()
    at = f" at {channel}mm" if channel else ""
    return f"nothing in the price list for {media}{at}"


def _suffix_area(suffix: str) -> Optional[float]:
    """The area a part-number suffix stands for: '020' -> 0.2, '1.8' -> 1.8."""
    suffix = (suffix or "").strip()
    if not suffix:
        return None
    try:
        return float(suffix) if "." in suffix else int(suffix) / 100.0
    except ValueError:
        return None


def quote_lines(items: Iterable[Dict[str, Any]],
                prices: Optional[Dict[str, float]] = None,
                rates: Optional[Dict[str, float]] = None) -> List[Dict[str, Any]]:
    """Price every line of an order. Unpriced lines are kept, marked and
    explained rather than dropped or quietly zeroed."""
    out: List[Dict[str, Any]] = []
    for it in items or []:
        try:
            qty = max(0, int(str(it.get("Quantity") or 0).strip()))
        except (TypeError, ValueError):
            qty = 0
        unit, source = price_for_item(it, prices, rates)
        out.append({
            "part_number": (it.get("Part Number") or "").strip().upper(),
            "description": describe_item(it),
            "quantity":    qty,
            "unit_price":  unit,
            "line_total":  round(unit * qty, 2),
            "source":      source,
            "priced_as":   it.get("_priced_as", "") if source == "near" else "",
            "reason":      "" if source else why_unpriced(it, prices),
            "item":        it,
        })
    return out


def source_label(line: Dict[str, Any]) -> str:
    """Where a line's price came from, as shown to whoever reads the quote."""
    source = line.get("source")
    if source == "list":
        return "Price list"
    if source == "rate":
        return "Rate per m²"
    if source == "near":
        return f"Priced as {line.get('priced_as') or 'nearest size'}"
    return line.get("reason") or "NOT PRICED"


def describe_item(item: Dict[str, Any]) -> str:
    """The line as it should read on a quote or an invoice.

    Plain ASCII throughout: this text goes into a Xero import as well as onto
    a quote, and an accounting import is the wrong place to find out how a
    piece of punctuation is decoded.
    """
    if item.get("item_kind") == "catalogue":
        # Whatever it was called in the price list, or what someone typed for
        # a one-off. It already reads the way it should on a quote.
        text = (str(item.get("Description") or "").strip()
                or str(item.get("Part Number") or "").strip()
                or "Item")
        kind = str(item.get("Product Type") or "").strip()
        if kind and kind.lower() not in text.lower():
            text = f"{kind} - {text}"
        return text

    if (item.get("item_kind") or "filter") == "bag":
        bits = [str(item.get("Filter Type") or "Bag Filter")]
        dims = " x ".join(str(item.get(k)) for k in ("width", "height")
                          if item.get(k))
        if dims:
            bits.append(dims + "mm")
        pockets = item.get("pockets")
        if pockets:
            bits.append(f"{pockets} pocket")
        return " - ".join(bits)

    bits = [str(item.get("Filter Type") or "Filter")]
    media = str(item.get("Media Type") or "").strip()
    if media:
        bits.append(media)
    s, l, c = item.get("Short"), item.get("Long"), item.get("Channel")
    if s and l:
        bits.append(f"{s} x {l} x {c or '?'}mm")
    area = _pn.nominal_square_metres(_pn.effective_area(item))
    if area > 0:
        bits.append(f"{area:.2f} m2")
    return " - ".join(bits)


def quote_totals(lines: Iterable[Dict[str, Any]],
                 gst_rate: float = 0.10) -> Dict[str, float]:
    """Subtotal, GST and total for a set of quote lines."""
    subtotal = round(sum(l.get("line_total", 0) for l in lines), 2)
    gst = round(subtotal * gst_rate, 2)
    return {"subtotal": subtotal, "gst": gst,
            "total": round(subtotal + gst, 2)}


def unpriced(lines: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """The lines nothing could price — what a quote must not go out with."""
    return [l for l in lines if not l.get("source")]


# ── Xero ─────────────────────────────────────────────────────────────────────
# Xero's own sales-invoice import format. Its column names and order are what
# Xero expects; changing them breaks the import, so they are written verbatim.

XERO_COLUMNS = [
    "*ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2",
    "POAddressLine3", "POAddressLine4", "POCity", "PORegion", "POPostalCode",
    "POCountry", "*InvoiceNumber", "Reference", "*InvoiceDate", "*DueDate",
    "InventoryItemCode", "*Description", "*Quantity", "*UnitAmount",
    "Discount", "*AccountCode", "*TaxType", "TrackingName1", "TrackingOption1",
    "TrackingName2", "TrackingOption2", "Currency", "BrandingTheme",
]

# Xero rejects an import outright if a date is in the wrong shape.
XERO_DATE = "%d/%m/%Y"


def _xero_date(value: str, fallback: Optional[_dt.date] = None) -> str:
    from .validation import parse_date
    dt = parse_date(value or "")
    if dt:
        return dt.strftime(XERO_DATE)
    return (fallback or _dt.date.today()).strftime(XERO_DATE)


def xero_rows(header: Dict[str, Any], lines: Iterable[Dict[str, Any]],
              customer: Optional[Dict[str, Any]] = None,
              account_code: str = "200",
              tax_type: str = "GST on Income",
              due_days: int = 30,
              invoice_number: str = "",
              reference: str = "") -> List[Dict[str, str]]:
    """One row per priced line, in Xero's sales-invoice import shape.

    Xero repeats the invoice-level fields on every line of the same invoice,
    which is why the contact and dates appear on each row.

    `invoice_number` and `reference` override the defaults. Invoicing a real
    order wants our own number on the invoice and the customer's purchase
    order in the reference, which is the opposite way round from a quote.
    """
    customer = customer or {}
    invoice_no = (invoice_number or "").strip() or (header.get("Order Number") or "").strip()
    issued = _xero_date(header.get("Date Ordered") or "")
    due_raw = (header.get("Date Due") or "").strip()
    if due_raw and due_raw.lower() != "asap":
        due = _xero_date(due_raw)
    else:
        issued_dt = _dt.datetime.strptime(issued, XERO_DATE).date()
        due = (issued_dt + _dt.timedelta(days=due_days)).strftime(XERO_DATE)

    contact = ((customer.get("legal_name") or "").strip()
               or (customer.get("name") or "").strip()
               or (header.get("Customer Name") or "").strip())

    rows: List[Dict[str, str]] = []
    for line in lines:
        if line.get("quantity", 0) <= 0:
            continue
        rows.append({
            "*ContactName":      contact,
            "EmailAddress":      (customer.get("email") or "").strip(),
            "POAddressLine1":    (customer.get("delivery_address1") or "").strip(),
            "POAddressLine2":    (customer.get("delivery_address2") or "").strip(),
            "POAddressLine3":    "",
            "POAddressLine4":    "",
            "POCity":            (customer.get("delivery_city") or "").strip(),
            "PORegion":          (customer.get("delivery_state") or "").strip(),
            "POPostalCode":      (customer.get("delivery_postcode") or "").strip(),
            "POCountry":         (customer.get("delivery_country") or "").strip(),
            "*InvoiceNumber":    invoice_no,
            "Reference":         (reference or "").strip() or (header.get("Job") or "").strip(),
            "*InvoiceDate":      issued,
            "*DueDate":          due,
            "InventoryItemCode": line.get("part_number", ""),
            "*Description":      line.get("description", ""),
            "*Quantity":         str(line.get("quantity", 0)),
            "*UnitAmount":       f'{line.get("unit_price", 0):.2f}',
            "Discount":          "",
            "*AccountCode":      account_code,
            "*TaxType":          tax_type,
            "TrackingName1":     "Region" if (header.get("Location") or "").strip() else "",
            "TrackingOption1":   (header.get("Location") or "").strip(),
            "TrackingName2":     "",
            "TrackingOption2":   "",
            "Currency":          "AUD",
            "BrandingTheme":     "",
        })
    return rows


def invoice_reference(header: Dict[str, Any], invoice_number: str = "") -> str:
    """What to put in Xero's Reference so the customer recognises the invoice.

    Their purchase order number is the thing they will look for, so it goes in
    the reference — unless it is already the invoice number, or it is one of
    ours (TAF-ON-…), which means them, not us, and would only be noise there.
    """
    order_no = (header.get("Order Number") or "").strip()
    job = (header.get("Job") or "").strip()
    from .part_numbers import is_supplied_order_number
    same = (invoice_number or "").strip().upper() == order_no.upper()
    if order_no and not same and not is_supplied_order_number(order_no):
        return f"PO {order_no}" + (f" - {job}" if job else "")
    return job


def invoice_for_order(header: Dict[str, Any],
                      items: Iterable[Dict[str, Any]],
                      customer: Optional[Dict[str, Any]] = None,
                      prices: Optional[Dict[str, float]] = None,
                      rates: Optional[Dict[str, float]] = None,
                      invoice_number: str = "",
                      account_code: str = "200",
                      tax_type: str = "GST on Income",
                      due_days: int = 30) -> Tuple[List[Dict[str, str]],
                                                   List[Dict[str, Any]]]:
    """Price a finished order and return (Xero rows, lines nothing could price).

    The unpriced lines come back rather than being dropped: an invoice that is
    quietly short a line is worse than one that doesn't go out, and the caller
    is the only thing that can ask a person what to do about it.
    """
    lines = quote_lines(items, prices, rates)
    number = (invoice_number or "").strip() or (header.get("Order Number") or "").strip()
    rows = xero_rows(header, [l for l in lines if l.get("source")],
                     customer=customer, account_code=account_code,
                     tax_type=tax_type, due_days=due_days,
                     invoice_number=number,
                     reference=invoice_reference(header, number))
    return rows, unpriced(lines)


def write_xero_csv(path, rows: Iterable[Dict[str, str]]) -> str:
    """Write Xero import rows to a CSV. Returns the path written."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=XERO_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)
    return str(path)
