import os
import subprocess
import platform
import atexit
import threading as _threading
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border
from datetime import datetime, timedelta

# ── Persistent Excel COM cache ────────────────────────────────────────────────
# Excel.Application is expensive to start (~3-8 s).  We keep one instance alive
# for the whole session and reuse it.  COM handles cross-thread marshaling
# transparently, so this is safe even when called from worker threads.
_excel_app  = None
_excel_lock = _threading.Lock()


def _ensure_excel_app():
    """Return a live Excel.Application COM object, creating one if needed."""
    global _excel_app
    with _excel_lock:
        if _excel_app is not None:
            try:
                _ = _excel_app.Version      # quick liveness check
                return _excel_app
            except Exception:
                _excel_app = None           # stale — will recreate below
        try:
            import win32com.client
            app = win32com.client.Dispatch("Excel.Application")
            app.Visible       = False
            app.DisplayAlerts = False
            _excel_app = app
        except Exception:
            _excel_app = None
        return _excel_app


def _quit_excel():
    """Called at process exit — gracefully close the cached Excel instance."""
    global _excel_app
    with _excel_lock:
        if _excel_app is not None:
            try:
                _excel_app.Quit()
            except Exception:
                pass
            _excel_app = None

atexit.register(_quit_excel)


def _export_pdf_via_cached_excel(xlsx_path: str, pdf_path: str) -> bool:
    """Export xlsx→PDF using the persistent Excel COM object.  Returns True on success."""
    for attempt in range(2):
        app = _ensure_excel_app()
        if app is None:
            return False
        try:
            wb = app.Workbooks.Open(xlsx_path)
            for ws in wb.Worksheets:
                ws.PageSetup.Orientation   = 2   # xlLandscape
                ws.PageSetup.Zoom          = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = 1
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
            wb.Close(SaveChanges=False)
            return os.path.exists(pdf_path)
        except Exception:
            # Excel may have become unresponsive — drop the reference and retry once
            with _excel_lock:
                _excel_app = None
            if attempt == 1:
                return False
    return False


# ── Cached template workbook ──────────────────────────────────────────────────
# load_workbook is ~0.5-1 s.  Cache the result so repeated orders skip it.
_template_wb       = None
_template_abs_path = None


def _load_template_cached():
    """Return the template workbook, loading it only once per process."""
    global _template_wb, _template_abs_path
    current = os.path.abspath(TEMPLATE_FILE)
    if _template_wb is None or _template_abs_path != current:
        _template_wb       = load_workbook(TEMPLATE_FILE)
        _template_abs_path = current
    return _template_wb


def prewarm_excel():
    """Call on a background thread at app startup to hide Excel's first-launch cost."""
    _ensure_excel_app()

def _export_pdf_powershell(xlsx_path: str, pdf_path: str) -> bool:
    """Export Excel to PDF via PowerShell COM automation. Returns True on success."""
    xlsx_ps = xlsx_path.replace("'", "''")
    pdf_ps  = pdf_path.replace("'", "''")
    script = (
        "$ErrorActionPreference='Stop';"
        "$xl=New-Object -ComObject Excel.Application;"
        "$xl.Visible=$false;"
        f"$wb=$xl.Workbooks.Open('{xlsx_ps}');"
        "foreach($ws in $wb.Worksheets){"
        "$ws.PageSetup.Orientation=2;"
        "$ws.PageSetup.Zoom=$false;"
        "$ws.PageSetup.FitToPagesWide=1;"
        "$ws.PageSetup.FitToPagesTall=1};"
        f"$wb.ExportAsFixedFormat(0,'{pdf_ps}');"
        "$wb.Close($false);"
        "$xl.Quit();"
        "[System.Runtime.InteropServices.Marshal]::ReleaseComObject($xl)|Out-Null"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True, timeout=90,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        return result.returncode == 0 and os.path.exists(pdf_path)
    except Exception:
        return False


def save_and_export_pdf(out_wb, out_path, auto_open=True):
    # Always save the .xlsx first — this never requires Excel
    out_wb.save(out_path)

    if platform.system() == "Windows":
        pdf_path = out_path.replace(".xlsx", ".pdf")

        # Method 1: persistent cached Excel COM (fast after first use)
        exported = _export_pdf_via_cached_excel(out_path, pdf_path)

        # Method 2: PowerShell COM fallback (works without pywin32)
        if not exported:
            exported = _export_pdf_powershell(out_path, pdf_path)

        if exported and os.path.exists(pdf_path):
            if auto_open:
                try:
                    os.startfile(pdf_path)
                except Exception:
                    pass
            return pdf_path

        # Excel not installed — xlsx already saved, just return it
        if auto_open:
            try:
                os.startfile(out_path)
            except Exception:
                # No app associated with .xlsx — silently continue;
                # the caller will show the file path to the user
                pass
        return out_path

    elif platform.system() == "Darwin":
        if auto_open:
            try:
                subprocess.call(["open", out_path])
            except Exception:
                pass
    else:
        if auto_open:
            try:
                subprocess.call(["xdg-open", out_path])
            except Exception:
                pass

    return out_path

TEMPLATE_FILE = "Templates.xlsx"
OUTPUT_DIR = "orders"

# Allowed stock sizes for stepped filters (short, long, channel)
_STOCK_SIZES = {(535, 535, 50), (535, 1135, 50)}

def _copy_values_and_styles(src_ws, dest_ws):
    """Copy values + basic styles from src_ws to dest_ws."""
    for row in src_ws.iter_rows():
        for src_cell in row:
            dest_cell = dest_ws[src_cell.coordinate]
            dest_cell.value = src_cell.value

            if src_cell.has_style:
                # Clone font
                dest_cell.font = Font(
                    name=src_cell.font.name,
                    size=src_cell.font.size,
                    bold=src_cell.font.bold,
                    italic=src_cell.font.italic,
                    underline=src_cell.font.underline,
                    color=src_cell.font.color
                )
                # Clone alignment
                dest_cell.alignment = Alignment(
                    horizontal=src_cell.alignment.horizontal,
                    vertical=src_cell.alignment.vertical,
                    wrap_text=src_cell.alignment.wrap_text,
                )
                # Clone border
                if src_cell.border:
                    dest_cell.border = Border(
                        left=src_cell.border.left,
                        right=src_cell.border.right,
                        top=src_cell.border.top,
                        bottom=src_cell.border.bottom
                    )


def _add_business_days(start_date: datetime, business_days: int) -> datetime:
    """Return date after adding business days (skipping weekends)."""
    current = start_date
    days_added = 0
    while days_added < business_days:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Mon-Fri are 0-4
            days_added += 1
    return current


def _format_date_due(date_ordered_raw, date_due_raw):
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            parsed = datetime.strptime(date_ordered_raw, fmt)
            break
        except ValueError:
            continue
    else:
        raise ValueError(f"Could not parse date ordered: {date_ordered_raw}")

    if date_due_raw and date_due_raw.strip().upper() != "ASAP":
        for fmt in ("%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                parsed_due = datetime.strptime(date_due_raw, fmt)
                break
            except ValueError:
                continue
        else:
            parsed_due = _add_business_days(parsed, 2)
    else:
        parsed_due = _add_business_days(parsed, 2)

    # Portable formatting
    date_ordered_out = f"{parsed.day}-{parsed.month}-{parsed.strftime('%y')}"  # 18-9-25
    date_due_out = parsed_due.strftime("%a %d-%m-%y").upper()                  # THU 18-9-25

    return date_ordered_out, date_due_out


def validate_line_item(item):
    short = int(item.get("Short", 0))
    long = int(item.get("Long", 0))
    channel = int(item.get("Channel", 0))
    ftype = item.get("Filter Type", "").lower()

    # Rule 1: Short <= Long
    if short > long:
        raise ValueError(f"Invalid dimensions: Short ({short}) > Long ({long}). Please swap or correct.")

    # Rule 2: Flat panel channel restriction
    if ftype == "flat panel":
        if not (channel <= 25 or channel == 50):
            raise ValueError(f"Flat Panel channel must be <=25 mm or exactly 50 mm (got {channel}).")

    return True


def _write_header_cells(ws, header_info, date_ordered_fmt, date_due_fmt):
    ws["C1"].value = header_info.get("Customer Name", "").upper()
    ws["C3"].value = header_info.get("Order Number", "")
    ws["G1"].value = date_ordered_fmt  # "DD-M-YY"
    ws["H3"].value = header_info.get("Attention", "").upper()
    # Job, Due, Location, Notes
    ws["G12"].value = f"JOB: {header_info.get('Job', '')}"
    ws["G14"].value = f"DUE: {date_due_fmt}"
    ws["G16"].value = f"LOCATION: {header_info.get('Location', '').upper()}"
    # G18 may be overwritten later for special notes; set base notes if present
    base_notes = header_info.get("Notes", "")
    ws["G18"].value = base_notes if base_notes else None


def _fill_markings(ws, short: int, long: int):
    """
    Fill steel marking cells according to Forward/Flipped/Reversed rules.
    Forward (G6:L6)
      Mark1 = 20
      Mark2 = Mark1 + short - 2
      Mark3 = Mark2 + long - 2
      Mark4 = Mark3 + short - 2
      Mark if cap = Mark4 + 20 
      Total Length = Mark4 + long - 2
    Flipped (G8:L8) swap short/long
    Reversed (G10:K10)
        Mark1 = Long - 2
        Mark2 = Mark1 + short - 2
        Mark3 = Mark2 + long - 2
        Mark4 = Mark3 + short - 2
        Mark if cap = Mark4 + 20
        No total length
    """
    # Forward
    m1 = 20
    m2 = m1 + short - 2
    m3 = m2 + long - 2
    m4 = m3 + short - 2
    mcap = m4 + 20
    mtot = m4 + long - 2
    forwardcap = long - 2
    flippedcap = short - 2
    ws["G6"].value = m1
    ws["H6"].value = m2
    ws["I6"].value = m3
    ws["J6"].value = m4
    ws["K6"].value = mcap
    ws["L6"].value = mtot
    ws["M6"].value = forwardcap

    # Flipped
    fm1 = 20
    fm2 = fm1 + long - 2
    fm3 = fm2 + short - 2
    fm4 = fm3 + long - 2
    fcap = fm4 + 20
    ftot = fm4 + short - 2
    ws["G8"].value = fm1
    ws["H8"].value = fm2
    ws["I8"].value = fm3
    ws["J8"].value = fm4
    ws["K8"].value = fcap
    ws["L8"].value = ftot
    ws["M8"].value = flippedcap

    # Reversed
    rm1 = long - 2
    rm2 = rm1 + short - 2
    rm3 = rm2 + long - 2
    rm4 = rm3 + short - 2
    rcap = rm4 + 20
    ws["G10"].value = rm1
    ws["H10"].value = rm2
    ws["I10"].value = rm3
    ws["J10"].value = rm4
    ws["K10"].value = rcap


def _clear_markings(ws):
    # Clear steel marking cells G6-M6, G8-M8, G10-K10
    cells = ["G6", "H6", "I6", "J6", "K6", "L6", "M6",
             "G8", "H8", "I8", "J8", "K8", "L8", "M8",
             "G10", "H10", "I10", "J10", "K10"]
    for c in cells:
        ws[c].value = None


def _clear_media_wire(ws):
    # Clear steel marking cells G6-M6, G8-M8, G10-K10
    cells = ["N6", "O6", "P6", "Q6"]
    for c in cells:
        ws[c].value = None


def _fill_vorf(ws, item, pleat=False, header_only=False, use_stock_v=False):
    """
    Fill the VorF (V-form / Flat panel / Pleat inserts / Header) sheet with calculations.
    - item: dict containing at least: Quantity, Short, Long, Channel, Filter Type, Media Type
    - pleat: if True, don't write steel markings
    - header_only: if True, don't write wire/media
    - use_stock_v: if True, skip markings/wire/media and add stock note
    """

    qty = int(item.get("Quantity", 0))
    short = int(item.get("Short", 0))
    long = int(item.get("Long", 0))
    channel = int(item.get("Channel", 0))
    ftype = item.get("Filter Type", "")
    media = item.get("Media Type", "")
    item_notes = item.get("Notes", "").strip()
    if item_notes:
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + " " + item_notes).strip()

    # Basic line item fields
    ws["B6"].value = media
    ws["C6"].value = qty

    # Determine label: V{channel} or F{channel}
    # For flat panel use F*, for v-form use V*, for pleat/header follow original selection logic
    if ftype.lower() == "flat panel":
        ws["D6"].value = f"F{channel}"
    elif ftype.lower() == "v-form":
        ws["D6"].value = f"V{channel}"
    else:
        # pleat/header could still be V or F depending on user's intended channel; try to infer:
        if str(item.get("Filter+Channel", "")).upper().startswith("F"):
            ws["D6"].value = item.get("Filter+Channel")
        else:
            ws["D6"].value = f"V{channel}"

    ws["E6"].value = short
    ws["F6"].value = long

    # Stock handling: if use_stock_v True, skip markings and wire/media but keep header+line info
    if use_stock_v:
        # Clear markings and wire/media cells explicitly
        _clear_markings(ws)
        _clear_media_wire(ws)
        # Add stock note (mention which stock size)
        note = f"STEPPED FILTER — USE STOCK V-FORM ({short}x{long}x{channel})"
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + " " + note).strip()
        return

    # Pleat inserts: skip steel markings entirely
    if pleat:
        _clear_markings(ws)
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + " PLEAT INSERTS ONLY").strip()
    # Header only: skip wire & media calculations
    elif header_only:
        _clear_media_wire(ws)
        _fill_markings(ws, short, long)
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + " HEADER ONLY").strip()
    else:
        # Write steel markings based on short/long
        _fill_markings(ws, short, long)

    # Now compute wire cuts (N6) and qty (O6), media cuts (P6) and qty (Q6)
    # Decide label to know whether Flat panel or V-form
    label = str(ws["D6"].value or "").upper()
    wire_short = short
    wire_long = long

    if label.startswith("F"):
        # Flat panels: -10 from short and long
        wire_short = short - 10
        wire_long = long - 10
    elif label.startswith("V"):
        # V-form: use depth (channel) bracket rules
        depth = channel
        note_over_100 = False
        if depth > 100:
            depth_use = 100  # use 85-100 formula
            note_over_100 = True
        else:
            depth_use = depth

        if 30 <= depth_use <= 60:
            wire_short = short - 15
            wire_long = int(round(long * 1.36))
        elif 65 <= depth_use <= 80:
            wire_short = short - 10
            wire_long = int(round(long * 1.55))
        elif 85 <= depth_use <= 100:
            wire_short = short - 10
            wire_long = int(round(long * 1.785))
        else:
            # fallback: leave as short/long unchanged
            wire_short = short
            wire_long = long

        if note_over_100:
            existing = ws["G18"].value or "NOTES:"
            ws["G18"].value = (existing + " CHECK WIRE & MEDIA – >100mm").strip()
    else:
        # unknown label - leave as is
        wire_short = short
        wire_long = long

    # Write wire cut and qty
    ws["N6"].value = f"{int(wire_short)}*{int(wire_long)}"
    ws["O6"].value = qty * 2

    # Media cut and qty
    if label.startswith("F"):
        media_short = short
        media_long = long
    else:
        media_short = int(wire_short) + 60
        media_long = int(wire_long) + 40

    ws["P6"].value = f"{int(media_short)}*{int(media_long)}"
    ws["Q6"].value = qty


def _fill_flyscreen(ws, item, stepped=False, use_stock_fs=False):
    """
    Fill flyscreen fields and calculations.
    - item contains Quantity, Short, Long, Media Type
    - stepped True indicates this flyscreen is part of a stepped filter
    - use_stock_fs True -> skip media calculations and add stock note
    """
    qty = int(item.get("Quantity", 0))
    short = int(item.get("Short", 0))
    long = int(item.get("Long", 0))
    media = item.get("Media Type", "")
    item_notes = item.get("Notes", "").strip()
    if item_notes:
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + " " + item_notes).strip()

    ws["B6"].value = "" if stepped and use_stock_fs else media
    ws["C6"].value = qty
    ws["D6"].value = "FS9"
    ws["E6"].value = short
    ws["F6"].value = long

    # Frame cuts and quantities
    ws["G6"].value = short - 2
    ws["H6"].value = qty * 2
    ws["I6"].value = long
    ws["J6"].value = qty * 2

    if use_stock_fs:
        # Skip media calculations but still keep header and line items
        ws["G6"].value = None
        ws["H6"].value = None
        ws["I6"].value = None
        ws["J6"].value = None  
        ws["K6"].value = None
        ws["L6"].value = None
        existing = ws["G18"].value or "NOTES:"
        ws["G18"].value = (existing + f" STEPPED FILTER — USE STOCK FLYSCREEN ({short}x{long}x50)").strip()
    else:
        # Media cut: +20 to short and long
        media_short = short + 20
        media_long = long + 20
        ws["K6"].value = f"{media_short}*{media_long}"
        ws["L6"].value = qty


def generate_order_workbook(header_data, items, auto_open=True,
                            page_start=1, grand_total=None):
    """
    Main entry point — generates a PDF directly using pdf_generator.
    No Excel installation required.
    """
    from pdf_generator import generate_pdf

    safe_c = "".join(c for c in (header_data.get("Customer Name") or "")
                     if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
    safe_n = "".join(c for c in (header_data.get("Order Number") or "")
                     if c.isalnum() or c in ("_", "-")).strip()
    base     = f"{safe_c}_{safe_n}".strip("_") or "order"
    pdf_path = os.path.join(os.getcwd(), f"{base}.pdf")

    # Validate items first (keeps existing validation)
    for item in items:
        try:
            validate_line_item(item)
        except Exception as exc:
            raise ValueError(str(exc))

    generate_pdf(header_data, items, pdf_path,
                 page_start=page_start, grand_total=grand_total)

    if auto_open and os.path.exists(pdf_path):
        try:
            os.startfile(pdf_path)
        except Exception:
            pass

    return pdf_path
